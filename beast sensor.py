import asyncio
import json
import math
import queue
import struct
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from bleak import BleakClient, BleakScanner
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


DEVICE_ADDRESS = "BE:A5:7F:30:78:68"
IMU_CHARACTERISTIC_UUID = "bea5760d-503d-4920-b000-101e7306b003"

# Detection settings for short bottom-to-top exercise movements.
CALIBRATION_SECONDS = 3.0
START_ACCELERATION_M_S2 = 0.10
START_CONFIRM_SECONDS = 0.05
ACCELERATION_FILTER_ALPHA = 0.35
ACCELERATION_HIGHPASS_CUTOFF_HZ = 0.18
IDLE_BASELINE_ALPHA = 0.01
MIN_ASCENT_SECONDS = 0.05
MIN_ASCENT_SPEED_M_S = 0.04
MIN_ASCENT_DISPLACEMENT_M = 0.008
MAX_ASCENT_SECONDS = 4.0
TOP_DECELERATION_M_S2 = 0.15
TOP_ZERO_SPEED_M_S = 0.015
TOP_STILL_ACCELERATION_M_S2 = 0.16
TOP_STILL_SECONDS = 0.08
TOP_REVERSAL_DISTANCE_M = 0.006
TOP_CONFIRM_SECONDS = 0.05
RETURN_MIN_SECONDS = 0.05
RETURN_MAX_SECONDS = 5.0
RETURN_DOWN_SPEED_M_S = 0.04
RETURN_ZERO_SPEED_M_S = 0.02
RETURN_STILL_ACCELERATION_M_S2 = 0.22
RETURN_STILL_SECONDS = 0.12
MAX_SAMPLE_INTERVAL_SECONDS = 0.10
MAX_VALID_DISPLACEMENT_M = 2.0
MAX_VALID_PEAK_SPEED_M_S = 4.0
GRAVITY_M_S2 = 9.80665

PROJECT_DIRECTORY = Path(__file__).resolve().parent
REPETITION_DATA_FILE = PROJECT_DIRECTORY / "beast_repetitions.json"
EXCEL_WORKBOOK = PROJECT_DIRECTORY / "outputs" / "beast_tracker" / "Beast Workout.xlsx"


@dataclass
class ImuSample:
    timestamp: float
    x_g: float
    y_g: float
    z_g: float
    vertical_g: float


@dataclass
class MotionSample:
    elapsed_s: float
    acceleration_m_s2: float
    raw_velocity_m_s: float
    raw_position_m: float


def decode_imu_packet(data: bytearray) -> ImuSample | None:
    """Decode quaternion and acceleration from a Beast IMU packet."""
    if len(data) < 16:
        return None

    q0, q1, q2, q3 = map(float, struct.unpack("<hhhh", data[2:10]))
    x_raw, y_raw, z_raw = struct.unpack("<hhh", data[10:16])
    quaternion_norm = math.sqrt(q0 * q0 + q1 * q1 + q2 * q2 + q3 * q3)
    if quaternion_norm == 0.0:
        return None

    q0 /= quaternion_norm
    q1 /= quaternion_norm
    q2 /= quaternion_norm
    q3 /= quaternion_norm

    x_g = x_raw / 1000.0
    y_g = y_raw / 1000.0
    z_g = z_raw / 1000.0
    vertical_g = (
        2.0 * (q1 * q2 + q3 * q0) * x_g
        + (1.0 - 2.0 * (q1 * q1 + q3 * q3)) * y_g
        + 2.0 * (q2 * q3 - q1 * q0) * z_g
    )
    return ImuSample(time.perf_counter(), x_g, y_g, z_g, vertical_g)


def corrected_ascent_metrics(samples: list[MotionSample]) -> dict[str, float]:
    """Enforce zero speed at the detected top and calculate ascent metrics."""
    duration_s = samples[-1].elapsed_s
    final_raw_velocity = samples[-1].raw_velocity_m_s
    previous_time = 0.0
    previous_velocity = 0.0
    position_m = 0.0
    min_position_m = 0.0
    max_position_m = 0.0
    peak_speed_m_s = 0.0

    for sample in samples:
        progress = sample.elapsed_s / duration_s if duration_s > 0.0 else 0.0
        velocity = max(
            0.0,
            sample.raw_velocity_m_s - final_raw_velocity * progress,
        )
        dt = sample.elapsed_s - previous_time
        position_m += 0.5 * (previous_velocity + velocity) * dt
        min_position_m = min(min_position_m, position_m)
        max_position_m = max(max_position_m, position_m)
        peak_speed_m_s = max(peak_speed_m_s, velocity)
        previous_time = sample.elapsed_s
        previous_velocity = velocity

    displacement_m = max_position_m - min_position_m
    return {
        "duration_s": duration_s,
        "displacement_m": displacement_m,
        "average_speed_m_s": displacement_m / duration_s if duration_s > 0.0 else 0.0,
        "peak_speed_m_s": peak_speed_m_s,
    }


def load_repetitions() -> list[dict]:
    if not REPETITION_DATA_FILE.exists():
        return []
    try:
        return json.loads(REPETITION_DATA_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def store_repetition(repetition: dict) -> None:
    repetitions = load_repetitions()
    repetitions.append(repetition)
    temporary_file = REPETITION_DATA_FILE.with_suffix(".json.tmp")
    temporary_file.write_text(json.dumps(repetitions, indent=2), encoding="utf-8")
    temporary_file.replace(REPETITION_DATA_FILE)


def build_excel_workbook() -> None:
    """Rebuild the portable Excel training log from saved repetitions."""
    repetitions = load_repetitions()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Repetitions"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A8"

    navy = "172554"
    blue = "2563EB"
    pale_blue = "DBEAFE"
    very_pale_blue = "EFF6FF"
    white = "FFFFFF"
    dark_blue = "1E3A8A"
    light_border = Side(style="thin", color="DCE6F1")

    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = "Beast Bar-Velocity Training Log"
    title.fill = PatternFill("solid", fgColor=navy)
    title.font = Font(bold=True, color=white, size=18)
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet["A3"] = "Completed repetitions"
    sheet["A4"] = "Average speed"
    sheet["A5"] = "Best peak speed"
    for row in range(3, 6):
        sheet[f"A{row}"].fill = PatternFill("solid", fgColor=pale_blue)
        sheet[f"A{row}"].font = Font(bold=True, color=dark_blue)
        sheet[f"B{row}"].fill = PatternFill("solid", fgColor=very_pale_blue)
        sheet[f"B{row}"].font = Font(bold=True, color=navy)

    sheet["B3"] = len(repetitions)
    sheet["B3"].number_format = "0"

    headers = [
        "Repetition",
        "Completed at",
        "Time (s)",
        "Displacement (m)",
        "Average speed (m/s)",
        "Peak speed (m/s)",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row=7, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_number, repetition in enumerate(repetitions, 8):
        completed_at = datetime.fromisoformat(repetition["completed_at"])
        if completed_at.tzinfo is not None:
            completed_at = completed_at.replace(tzinfo=None)
        values = [
            repetition["rep"],
            completed_at,
            repetition["duration_s"],
            repetition["displacement_m"],
            repetition["average_speed_m_s"],
            repetition["peak_speed_m_s"],
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.border = Border(bottom=light_border)
        sheet.cell(row=row_number, column=1).number_format = "0"
        sheet.cell(row=row_number, column=2).number_format = "yyyy-mm-dd hh:mm:ss"
        for column in range(3, 7):
            sheet.cell(row=row_number, column=column).number_format = "0.000"

    if repetitions:
        last_row = 7 + len(repetitions)
        sheet["B4"] = f"=IFERROR(AVERAGE(E8:E{last_row}),0)"
        sheet["B5"] = f"=IFERROR(MAX(F8:F{last_row}),0)"
        table = Table(displayName="RepetitionsTable", ref=f"A7:F{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet["B4"] = 0
        sheet["B5"] = 0

    sheet["B4"].number_format = '0.000 "m/s"'
    sheet["B5"].number_format = '0.000 "m/s"'
    widths = {"A": 22, "B": 22, "C": 16, "D": 20, "E": 22, "F": 20}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(3, max(8, 8 + len(repetitions))):
        sheet.row_dimensions[row].height = 20

    EXCEL_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    temporary_workbook = EXCEL_WORKBOOK.with_name("Beast Workout.tmp.xlsx")
    workbook.save(temporary_workbook)
    temporary_workbook.replace(EXCEL_WORKBOOK)

class ExcelWorkbookUpdater:
    """Serialize background workbook rebuilds so fast repetitions cannot race."""

    def __init__(self) -> None:
        self.requests: queue.Queue[bool | None] = queue.Queue()
        self.last_error: str | None = None
        self.worker = threading.Thread(target=self._run, daemon=False)
        self.worker.start()

    def request_update(self) -> None:
        self.requests.put(True)

    def close(self) -> None:
        self.requests.put(None)
        self.worker.join(timeout=45.0)

    def _run(self) -> None:
        while True:
            request = self.requests.get()
            if request is None:
                return

            stop_after_update = False
            while True:
                try:
                    next_request = self.requests.get_nowait()
                except queue.Empty:
                    break
                if next_request is None:
                    stop_after_update = True
                    break

            try:
                build_excel_workbook()
                self.last_error = None
            except Exception as exc:
                self.last_error = str(exc)
                print(f"Excel update failed: {exc}")

            if stop_after_update:
                return

class AscentRepTracker:
    WAITING = "waiting"
    ASCENDING = "ascending"
    RETURNING = "returning"

    def __init__(self, workbook_updater: ExcelWorkbookUpdater) -> None:
        self.workbook_updater = workbook_updater
        existing_repetitions = load_repetitions()
        self.rep_number = max(
            (int(rep.get("rep", 0)) for rep in existing_repetitions),
            default=0,
        )
        self.reset_connection()

    def reset_connection(self) -> None:
        self.state = self.WAITING
        self.calibration_started: float | None = None
        self.calibration_values: list[float] = []
        self.last_countdown: int | None = None
        self.gravity_baseline_g: float | None = None
        self.last_timestamp: float | None = None
        self.lowpass_acceleration = 0.0
        self.previous_lowpass_acceleration = 0.0
        self.filtered_acceleration = 0.0
        self.previous_acceleration = 0.0
        self.start_candidate: float | None = None
        self.start_candidate_values: list[float] = []
        self.upward_sign: float | None = None
        self.ascent_started = 0.0
        self.velocity_m_s = 0.0
        self.position_m = 0.0
        self.peak_velocity_m_s = 0.0
        self.peak_position_m = 0.0
        self.peak_index = 0
        self.top_candidate: float | None = None
        self.top_candidate_index = 0
        self.deceleration_seen = False
        self.top_still_candidate: float | None = None
        self.motion_samples: list[MotionSample] = []
        self.return_started = 0.0
        self.return_velocity_m_s = 0.0
        self.return_position_m = 0.0
        self.return_previous_acceleration = 0.0
        self.return_has_descended = False
        self.return_still_candidate: float | None = None

    def handle_packet(self, _sender, data: bytearray) -> None:
        sample = decode_imu_packet(data)
        if sample is None:
            return

        if self.gravity_baseline_g is None:
            self._calibrate(sample)
            return

        if self.last_timestamp is None:
            self.last_timestamp = sample.timestamp
            return

        dt = sample.timestamp - self.last_timestamp
        self.last_timestamp = sample.timestamp
        if dt <= 0.0 or dt > MAX_SAMPLE_INTERVAL_SECONDS:
            self.previous_acceleration = 0.0
            return

        raw_acceleration = (sample.vertical_g - self.gravity_baseline_g) * GRAVITY_M_S2
        self.lowpass_acceleration = (
            ACCELERATION_FILTER_ALPHA * raw_acceleration
            + (1.0 - ACCELERATION_FILTER_ALPHA) * self.lowpass_acceleration
        )
        highpass_tau = 1.0 / (2.0 * math.pi * ACCELERATION_HIGHPASS_CUTOFF_HZ)
        highpass_alpha = highpass_tau / (highpass_tau + dt)
        self.filtered_acceleration = highpass_alpha * (
            self.filtered_acceleration
            + self.lowpass_acceleration
            - self.previous_lowpass_acceleration
        )
        self.previous_lowpass_acceleration = self.lowpass_acceleration

        if self.state == self.WAITING:
            self._wait_for_ascent(sample)
        elif self.state == self.ASCENDING:
            self._track_ascent(sample, dt)
        else:
            self._track_return(sample, dt)

        self.previous_acceleration = self.filtered_acceleration

    def _calibrate(self, sample: ImuSample) -> None:
        if self.calibration_started is None:
            self.calibration_started = sample.timestamp

        self.calibration_values.append(sample.vertical_g)
        elapsed = sample.timestamp - self.calibration_started
        remaining = max(0, math.ceil(CALIBRATION_SECONDS - elapsed))
        if remaining > 0 and remaining != self.last_countdown:
            print(f"Calibration: {remaining}")
            self.last_countdown = remaining

        if elapsed >= CALIBRATION_SECONDS:
            self.gravity_baseline_g = sum(self.calibration_values) / len(self.calibration_values)
            self.last_timestamp = sample.timestamp
            print("Ready. Lift the bar from the bottom position.\n")

    def _directed_acceleration(self) -> float:
        if self.upward_sign is None:
            return abs(self.filtered_acceleration)
        return self.upward_sign * self.filtered_acceleration

    def _wait_for_ascent(self, sample: ImuSample) -> None:
        self.gravity_baseline_g = (
            (1.0 - IDLE_BASELINE_ALPHA) * self.gravity_baseline_g
            + IDLE_BASELINE_ALPHA * sample.vertical_g
        )
        start_signal = self._directed_acceleration()

        if start_signal >= START_ACCELERATION_M_S2:
            if self.start_candidate is None:
                self.start_candidate = sample.timestamp
                self.start_candidate_values = [self.filtered_acceleration]
            else:
                self.start_candidate_values.append(self.filtered_acceleration)
                if sample.timestamp - self.start_candidate >= START_CONFIRM_SECONDS:
                    if self.upward_sign is None:
                        average = sum(self.start_candidate_values) / len(self.start_candidate_values)
                        self.upward_sign = 1.0 if average >= 0.0 else -1.0
                    self._start_ascent(sample)
        else:
            self.start_candidate = None
            self.start_candidate_values = []

    def _start_ascent(self, sample: ImuSample) -> None:
        self.state = self.ASCENDING
        self.ascent_started = sample.timestamp
        self.velocity_m_s = 0.0
        self.position_m = 0.0
        self.peak_velocity_m_s = 0.0
        self.peak_position_m = 0.0
        self.peak_index = 0
        self.top_candidate = None
        self.deceleration_seen = False
        self.top_still_candidate = None
        directed_acceleration = self.upward_sign * self.filtered_acceleration
        self.motion_samples = [MotionSample(0.0, directed_acceleration, 0.0, 0.0)]
        self.start_candidate = None
        self.start_candidate_values = []

    def _track_ascent(self, sample: ImuSample, dt: float) -> None:
        acceleration = self.upward_sign * self.filtered_acceleration
        previous_velocity = self.velocity_m_s
        self.velocity_m_s += 0.5 * (
            self.upward_sign * self.previous_acceleration + acceleration
        ) * dt
        self.position_m += 0.5 * (previous_velocity + self.velocity_m_s) * dt
        elapsed = sample.timestamp - self.ascent_started
        self.motion_samples.append(
            MotionSample(elapsed, acceleration, self.velocity_m_s, self.position_m)
        )

        self.peak_velocity_m_s = max(self.peak_velocity_m_s, self.velocity_m_s)
        if self.position_m > self.peak_position_m:
            self.peak_position_m = self.position_m
            self.peak_index = len(self.motion_samples) - 1

        if elapsed >= MAX_ASCENT_SECONDS:
            self._discard_ascent(sample)
            return

        ascent_established = (
            elapsed >= MIN_ASCENT_SECONDS
            and self.peak_velocity_m_s >= MIN_ASCENT_SPEED_M_S
            and self.peak_position_m >= MIN_ASCENT_DISPLACEMENT_M
        )
        if not ascent_established:
            return

        if acceleration <= -TOP_DECELERATION_M_S2:
            self.deceleration_seen = True

        velocity_reversed = (
            self.deceleration_seen and self.velocity_m_s <= TOP_ZERO_SPEED_M_S
        )
        position_reversed = (
            self.deceleration_seen
            and self.position_m <= self.peak_position_m - TOP_REVERSAL_DISTANCE_M
        )
        if velocity_reversed or position_reversed:
            if self.top_candidate is None:
                self.top_candidate = sample.timestamp
                self.top_candidate_index = self.peak_index
            elif sample.timestamp - self.top_candidate >= TOP_CONFIRM_SECONDS:
                self._finish_ascent(self.top_candidate_index, sample)
                return
        else:
            self.top_candidate = None

        if self.deceleration_seen and abs(acceleration) <= TOP_STILL_ACCELERATION_M_S2:
            if self.top_still_candidate is None:
                self.top_still_candidate = sample.timestamp
            elif sample.timestamp - self.top_still_candidate >= TOP_STILL_SECONDS:
                self._finish_ascent(len(self.motion_samples) - 1, sample)
        else:
            self.top_still_candidate = None
    def _metrics_are_physical(self, metrics: dict[str, float]) -> bool:
        return (
            MIN_ASCENT_SECONDS <= metrics["duration_s"] <= MAX_ASCENT_SECONDS
            and MIN_ASCENT_DISPLACEMENT_M
            <= metrics["displacement_m"]
            <= MAX_VALID_DISPLACEMENT_M
            and 0.0 < metrics["average_speed_m_s"] <= MAX_VALID_PEAK_SPEED_M_S
            and 0.0 < metrics["peak_speed_m_s"] <= MAX_VALID_PEAK_SPEED_M_S
        )

    def _finish_ascent(self, top_index: int, sample: ImuSample) -> None:
        top_index = max(1, min(top_index, len(self.motion_samples) - 1))
        ascent_samples = self.motion_samples[: top_index + 1]
        metrics = corrected_ascent_metrics(ascent_samples)

        if self._metrics_are_physical(metrics):
            self.rep_number += 1
            repetition = {
                "rep": self.rep_number,
                "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                **metrics,
            }
            store_repetition(repetition)
            self.workbook_updater.request_update()
            print(
                f"REP {self.rep_number:02d} | "
                f"time {metrics['duration_s']:.2f} s | "
                f"displacement {metrics['displacement_m']:.3f} m | "
                f"average speed {metrics['average_speed_m_s']:.3f} m/s | "
                f"peak speed {metrics['peak_speed_m_s']:.3f} m/s"
            )

        self._enter_returning(sample)

    def _discard_ascent(self, sample: ImuSample) -> None:
        self._enter_returning(sample)

    def _enter_returning(self, sample: ImuSample) -> None:
        self.state = self.RETURNING
        self.return_started = sample.timestamp
        self.return_velocity_m_s = 0.0
        self.return_position_m = 0.0
        self.return_previous_acceleration = self.upward_sign * self.filtered_acceleration
        self.return_has_descended = False
        self.return_still_candidate = None
        self.motion_samples = []
        self.start_candidate = None
        self.start_candidate_values = []

    def _track_return(self, sample: ImuSample, dt: float) -> None:
        acceleration = self.upward_sign * self.filtered_acceleration
        previous_velocity = self.return_velocity_m_s
        self.return_velocity_m_s += 0.5 * (
            self.return_previous_acceleration + acceleration
        ) * dt
        self.return_position_m += 0.5 * (
            previous_velocity + self.return_velocity_m_s
        ) * dt
        self.return_previous_acceleration = acceleration
        elapsed = sample.timestamp - self.return_started

        if (
            self.return_velocity_m_s <= -RETURN_DOWN_SPEED_M_S
            or acceleration <= -TOP_DECELERATION_M_S2
        ):
            self.return_has_descended = True

        bottom_reversal = (
            self.return_has_descended
            and self.return_velocity_m_s >= -RETURN_ZERO_SPEED_M_S
            and acceleration >= START_ACCELERATION_M_S2
        )
        if bottom_reversal and elapsed >= RETURN_MIN_SECONDS:
            if self.start_candidate is None:
                self.start_candidate = sample.timestamp
                self.start_candidate_values = [self.filtered_acceleration]
            else:
                self.start_candidate_values.append(self.filtered_acceleration)
                if sample.timestamp - self.start_candidate >= START_CONFIRM_SECONDS:
                    self._start_ascent(sample)
                    return
        else:
            self.start_candidate = None
            self.start_candidate_values = []

        if (
            elapsed >= RETURN_MIN_SECONDS
            and abs(acceleration) <= RETURN_STILL_ACCELERATION_M_S2
        ):
            if self.return_still_candidate is None:
                self.return_still_candidate = sample.timestamp
            elif sample.timestamp - self.return_still_candidate >= RETURN_STILL_SECONDS:
                self._ready_at_bottom(sample)
                return
        else:
            self.return_still_candidate = None

        if elapsed >= RETURN_MAX_SECONDS:
            self._ready_at_bottom(sample)
    def _ready_at_bottom(self, sample: ImuSample) -> None:
        self.state = self.WAITING
        self.last_timestamp = sample.timestamp
        self.lowpass_acceleration = 0.0
        self.previous_lowpass_acceleration = 0.0
        self.filtered_acceleration = 0.0
        self.previous_acceleration = 0.0
        self.start_candidate = None
        self.start_candidate_values = []

async def scan() -> None:
    """Optional BLE discovery helper."""
    devices = await BleakScanner.discover(timeout=15.0)
    for device in devices:
        print(f"Name: {device.name or 'Unknown'} | Address: {device.address}")


async def list_services(address: str) -> None:
    """Optional GATT service-listing helper."""
    async with BleakClient(address) as client:
        for service in client.services:
            print(f"Service: {service.uuid} - {service.description}")
            for characteristic in service.characteristics:
                print(
                    f"  Characteristic: {characteristic.uuid} | "
                    f"props={','.join(characteristic.properties)}"
                )


async def stream_sensor(tracker: AscentRepTracker) -> None:
    disconnected = asyncio.Event()

    def disconnected_callback(_client: BleakClient) -> None:
        disconnected.set()

    async with BleakClient(
        DEVICE_ADDRESS,
        disconnected_callback=disconnected_callback,
    ) as client:
        characteristic = client.services.get_characteristic(IMU_CHARACTERISTIC_UUID)
        if not characteristic or "notify" not in characteristic.properties:
            raise RuntimeError("The Beast IMU notification characteristic is unavailable.")

        print(f"Connected to Beast. Excel: {EXCEL_WORKBOOK}\n")
        tracker.reset_connection()
        await client.start_notify(IMU_CHARACTERISTIC_UUID, tracker.handle_packet)
        await disconnected.wait()


async def main() -> None:
    # Optional discovery/debugging. Uncomment only when needed:
    # await scan()
    # await list_services(DEVICE_ADDRESS)

    workbook_updater = ExcelWorkbookUpdater()
    if not EXCEL_WORKBOOK.exists():
        workbook_updater.request_update()
    tracker = AscentRepTracker(workbook_updater)

    try:
        while True:
            try:
                await stream_sensor(tracker)
                print("Beast disconnected. Reconnecting...")
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                print(f"Connection error: {exc}. Retrying...")
            await asyncio.sleep(2.0)
    finally:
        workbook_updater.close()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nStopped by user.")